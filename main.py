"""
Nathan Code List - Centre de commandement freelance
Backend Python avec pywebview : vraie fenêtre desktop native (WebView2).

Compatible exécution directe (`python main.py`) et exécutable PyInstaller
(--onefile). Le fichier data.json est stocké dans %APPDATA%/NathanCodeList
pour persister entre les mises à jour de l'application.
"""

import json
import os
import sys
import uuid
from datetime import datetime, date

import webview

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import caldav
    from caldav.lib.error import AuthorizationError, NotFoundError
except ImportError:
    caldav = None
    AuthorizationError = Exception
    NotFoundError = Exception


# ----------------------------------------------------------------------
# Résolution des chemins (dev + PyInstaller --onefile)
# ----------------------------------------------------------------------
def _bundle_root() -> str:
    """Racine des assets : dossier du script en dev, _MEIPASS en bundle."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS  # type: ignore[attr-defined]
    return os.path.dirname(os.path.abspath(__file__))


def _user_data_dir() -> str:
    """Dossier utilisateur où data.json est conservé entre les sessions."""
    appdata = os.environ.get("APPDATA") or os.path.expanduser("~")
    target = os.path.join(appdata, "NathanCodeList")
    os.makedirs(target, exist_ok=True)
    return target


BUNDLE_DIR = _bundle_root()
INDEX_FILE = os.path.join(BUNDLE_DIR, "web", "index.html")
DATA_FILE = os.path.join(_user_data_dir(), "data.json")

DEFAULT_DATA = {
    "tasks": [],
    "prospection": {
        "sent": 0,
        "viewed": 0,
        "responses": 0,
        "meetings": 0,
    },
    "productivity": {
        "codeTime": 0,
        "prospectionTime": 0,
        "dailyGoalCode": 14400,
        "dailyGoalProspection": 3600,
        "lastReset": str(date.today()),
    },
    "projects": [],
    "events": [],
    "history": [],
    "icloud": {
        "apple_id": "",
        "app_password": "",
        "calendar_name": "Nathan Code List",
        "last_sync": "",
        "pending_deletes": [],
    },
}

ICLOUD_CALDAV_URL = "https://caldav.icloud.com/"


# ----------------------------------------------------------------------
# Import Excel : matrice "Import Calendrier"
# ----------------------------------------------------------------------
_EXCEL_COLS = {
    "calendar": 2,     # B
    "title": 3,        # C
    "start": 4,        # D
    "end": 5,          # E
    "location": 6,     # F
    "reminder": 7,     # G
    "description": 8,  # H
    "recurrence": 9,   # I (optionnelle : Jamais / Tous les jours / etc.)
}


def _normalize_recurrence(value) -> str:
    """Accepte 'Jamais', 'tous les jours', 'weekly', etc. → label FR canonique."""
    if value is None:
        return "Jamais"
    s = str(value).strip().lower()
    if not s or s in ("jamais", "aucune", "none", ""):
        return "Jamais"
    if "jour" in s or s in ("daily", "day"):
        return "Tous les jours"
    if "2 sem" in s or "deux sem" in s or "bi-weekly" in s or "biweekly" in s:
        return "Toutes les 2 semaines"
    if "sem" in s or s in ("weekly", "week"):
        return "Toutes les semaines"
    if "mois" in s or s in ("monthly", "month"):
        return "Tous les mois"
    if "an" in s or s in ("yearly", "year", "annual"):
        return "Tous les ans"
    return "Jamais"


def _to_iso(value) -> str:
    """Accepte datetime, date ou str ('2026-05-18 10:00') → ISO."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    # string : on tente plusieurs formats avant de retomber sur la valeur brute
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M",
                "%d/%m/%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).isoformat(timespec="minutes")
        except ValueError:
            continue
    return s


def _ics_escape(text: str) -> str:
    """Échappe les caractères spéciaux pour iCalendar."""
    if not text:
        return ""
    return (str(text)
            .replace("\\", "\\\\")
            .replace(";", "\\;")
            .replace(",", "\\,")
            .replace("\n", "\\n"))


def _ics_dt(iso: str) -> str:
    """Convertit ISO ('2026-05-18T10:00') en DTSTART iCal flottant local."""
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace(" ", "T"))
        return dt.strftime("%Y%m%dT%H%M%S")
    except ValueError:
        return ""


_REMINDER_MINUTES = {
    "à l'heure de l'événement": 0,
    "5min avant": 5,
    "15min avant": 15,
    "30 min avant": 30,
    "1h avant": 60,
    "1 jour avant": 1440,
}

# Mapping libellé FR → composantes RRULE iCalendar
_RECURRENCE_RRULE = {
    "Jamais": None,
    "Tous les jours": "FREQ=DAILY",
    "Toutes les semaines": "FREQ=WEEKLY",
    "Toutes les 2 semaines": "FREQ=WEEKLY;INTERVAL=2",
    "Tous les mois": "FREQ=MONTHLY",
    "Tous les ans": "FREQ=YEARLY",
}
_RRULE_TO_LABEL = {v: k for k, v in _RECURRENCE_RRULE.items() if v}


def _event_to_ics(ev: dict) -> str:
    """Sérialise un event en VEVENT autonome (VCALENDAR enveloppant)."""
    dtstart = _ics_dt(ev.get("start", ""))
    dtend = _ics_dt(ev.get("end", "")) or dtstart
    if not dtstart:
        return ""
    uid = f"{ev.get('id', uuid.uuid4().hex)}@nathancodelist"
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Nathan Code List//FR",
        "CALSCALE:GREGORIAN",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{now}",
        f"DTSTART:{dtstart}",
        f"DTEND:{dtend}",
        f"SUMMARY:{_ics_escape(ev.get('title', ''))}",
    ]
    if ev.get("location"):
        lines.append(f"LOCATION:{_ics_escape(ev['location'])}")
    if ev.get("description"):
        lines.append(f"DESCRIPTION:{_ics_escape(ev['description'])}")
    cat = ev.get("calendar")
    if cat:
        lines.append(f"CATEGORIES:{_ics_escape(cat)}")

    # Récurrence : label FR → RRULE iCal
    rec_label = ev.get("recurrence", "Jamais")
    rrule = _RECURRENCE_RRULE.get(rec_label)
    if rrule:
        lines.append(f"RRULE:{rrule}")

    reminder = ev.get("reminder", "")
    if reminder in _REMINDER_MINUTES:
        minutes = _REMINDER_MINUTES[reminder]
        lines += [
            "BEGIN:VALARM",
            "ACTION:DISPLAY",
            f"DESCRIPTION:{_ics_escape(ev.get('title', ''))}",
            # PT0M = à l'heure pile de l'événement ; sinon X min avant.
            f"TRIGGER:-PT{minutes}M" if minutes > 0 else "TRIGGER:PT0M",
            "END:VALARM",
        ]

    lines += ["END:VEVENT", "END:VCALENDAR"]
    return "\r\n".join(lines)


def _icaldt_to_iso(value) -> str:
    """Convertit un dt iCal (date ou datetime, tz-aware ou non) en ISO local."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone().replace(tzinfo=None)
        return value.isoformat(timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _ics_to_event(ics_text: str, fallback_calendar: str = "iCloud") -> dict | None:
    """Parse une string ICS en event dict Nathan Code List."""
    try:
        from icalendar import Calendar as ICal
    except ImportError:
        return None
    try:
        cal = ICal.from_ical(ics_text)
    except Exception:
        return None
    for comp in cal.walk("VEVENT"):
        uid = str(comp.get("UID", "")) or str(uuid.uuid4())
        summary = str(comp.get("SUMMARY", "")) or "(Sans titre)"
        location = str(comp.get("LOCATION", "")) if comp.get("LOCATION") else ""
        description = (str(comp.get("DESCRIPTION", ""))
                       if comp.get("DESCRIPTION") else "")
        categories = comp.get("CATEGORIES")
        cat = fallback_calendar
        if categories is not None:
            try:
                if hasattr(categories, "cats") and categories.cats:
                    cat = str(categories.cats[0])
                else:
                    cat = str(categories).split(",")[0].strip() or fallback_calendar
            except Exception:
                cat = fallback_calendar
        start = _icaldt_to_iso(getattr(comp.get("DTSTART"), "dt", None))
        end = _icaldt_to_iso(getattr(comp.get("DTEND"), "dt", None))

        # RRULE → label FR
        recurrence = "Jamais"
        rrule = comp.get("RRULE")
        if rrule is not None:
            try:
                parts = []
                for k, v in rrule.items():
                    val = v[0] if isinstance(v, list) and v else v
                    parts.append(f"{k}={val}")
                key = ";".join(parts)
                recurrence = _RRULE_TO_LABEL.get(key, "Jamais")
                # fallback : si on ne reconnaît pas, on déduit du FREQ seul
                if recurrence == "Jamais":
                    freq = rrule.get("FREQ")
                    if freq:
                        freq_val = freq[0] if isinstance(freq, list) else freq
                        recurrence = _RRULE_TO_LABEL.get(
                            f"FREQ={freq_val}", "Jamais")
            except Exception:
                pass

        # VALARM → label FR
        reminder = "aucun"
        for sub in comp.subcomponents:
            if getattr(sub, "name", "") == "VALARM":
                trigger = sub.get("TRIGGER")
                if trigger is None:
                    continue
                td = getattr(trigger, "dt", None)
                if td is None:
                    continue
                # td est un timedelta négatif (avant l'event) ou 0
                if hasattr(td, "total_seconds"):
                    mins = int(round(-td.total_seconds() / 60))
                else:
                    mins = 0
                if mins <= 0:
                    reminder = "à l'heure de l'événement"
                elif mins == 5:
                    reminder = "5min avant"
                elif mins == 15:
                    reminder = "15min avant"
                elif mins == 30:
                    reminder = "30 min avant"
                elif mins == 60:
                    reminder = "1h avant"
                elif mins == 1440:
                    reminder = "1 jour avant"
                break

        return {
            "uid": uid,
            "calendar": cat,
            "title": summary,
            "start": start,
            "end": end,
            "location": location,
            "reminder": reminder,
            "recurrence": recurrence,
            "description": description,
        }
    return None


def _parse_events_from_xlsx(path: str) -> list[dict]:
    """Lit la première feuille au format 'Matrice d'Import Calendrier'.
    Cherche la ligne d'en-tête (contient 'Calendrier' en col B) puis lit
    toutes les lignes en-dessous tant qu'il y a un titre."""
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl n'est pas installé. Lance : pip install openpyxl"
        )
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        cell = ws.cell(row=r, column=_EXCEL_COLS["calendar"]).value
        if isinstance(cell, str) and cell.strip().lower() == "calendrier":
            header_row = r
            break
    if header_row is None:
        header_row = 4  # défaut conforme au template fourni

    events: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        title = ws.cell(row=r, column=_EXCEL_COLS["title"]).value
        if not title:
            continue
        events.append({
            "id": str(uuid.uuid4()),
            "calendar": (ws.cell(row=r, column=_EXCEL_COLS["calendar"]).value
                         or "Travail"),
            "title": str(title).strip(),
            "start": _to_iso(ws.cell(row=r, column=_EXCEL_COLS["start"]).value),
            "end": _to_iso(ws.cell(row=r, column=_EXCEL_COLS["end"]).value),
            "location": ws.cell(row=r, column=_EXCEL_COLS["location"]).value or "",
            "reminder": ws.cell(row=r, column=_EXCEL_COLS["reminder"]).value or "",
            "description": ws.cell(row=r, column=_EXCEL_COLS["description"]).value or "",
            "recurrence": _normalize_recurrence(
                ws.cell(row=r, column=_EXCEL_COLS["recurrence"]).value),
            "createdAt": datetime.now().isoformat(timespec="seconds"),
            "source": "excel",
        })
    return events


# ----------------------------------------------------------------------
# Persistance JSON
# ----------------------------------------------------------------------
def _ensure_data_file() -> None:
    if not os.path.exists(DATA_FILE):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_DATA, f, indent=2, ensure_ascii=False)


def _read() -> dict:
    _ensure_data_file()
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Fusion défensive
    for key, value in DEFAULT_DATA.items():
        if key not in data:
            data[key] = value
        elif isinstance(value, dict):
            for sub_k, sub_v in value.items():
                if sub_k not in data[key]:
                    data[key][sub_k] = sub_v

    # Reset journalier
    today = str(date.today())
    if data["productivity"].get("lastReset") != today:
        data["history"].append({
            "date": data["productivity"].get("lastReset", today),
            "codeTime": data["productivity"]["codeTime"],
            "prospectionTime": data["productivity"]["prospectionTime"],
        })
        data["productivity"]["codeTime"] = 0
        data["productivity"]["prospectionTime"] = 0
        data["productivity"]["lastReset"] = today
        _write(data)

    return data


def _write(data: dict) -> None:
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ----------------------------------------------------------------------
# API exposée à React via pywebview (window.pywebview.api.*)
# ----------------------------------------------------------------------
class Api:
    # ---------- Lecture / écriture brute ----------
    def get_data(self) -> dict:
        return _read()

    def save_data(self, data: dict) -> dict:
        _write(data)
        return data

    # ---------- Tâches ----------
    def add_task(self, title: str, priority: str = "medium") -> dict:
        data = _read()
        data["tasks"].append({
            "id": str(uuid.uuid4()),
            "title": (title or "").strip(),
            "priority": priority,
            "completed": False,
            "timeSpent": 0,
            "createdAt": datetime.now().isoformat(timespec="seconds"),
        })
        _write(data)
        return data

    def update_task(self, task_id: str, patch: dict) -> dict:
        data = _read()
        for t in data["tasks"]:
            if t["id"] == task_id:
                t.update(patch)
                break
        _write(data)
        return data

    def add_task_time(self, task_id: str, seconds: int) -> dict:
        data = _read()
        for t in data["tasks"]:
            if t["id"] == task_id:
                t["timeSpent"] = t.get("timeSpent", 0) + int(seconds)
                break
        data["productivity"]["codeTime"] = (
            data["productivity"].get("codeTime", 0) + int(seconds)
        )
        _write(data)
        return data

    def delete_task(self, task_id: str) -> dict:
        data = _read()
        data["tasks"] = [t for t in data["tasks"] if t["id"] != task_id]
        _write(data)
        return data

    # ---------- Prospection ----------
    def update_prospection(self, patch: dict) -> dict:
        data = _read()
        data["prospection"].update(patch)
        _write(data)
        return data

    def add_prospection_time(self, seconds: int) -> dict:
        data = _read()
        data["productivity"]["prospectionTime"] = (
            data["productivity"].get("prospectionTime", 0) + int(seconds)
        )
        _write(data)
        return data

    # ---------- Projets ----------
    def add_project(self, name: str, client: str = "",
                    status: str = "prospect", description: str = "") -> dict:
        data = _read()
        data["projects"].append({
            "id": str(uuid.uuid4()),
            "name": (name or "").strip(),
            "client": (client or "").strip(),
            "status": status,
            "description": (description or "").strip(),
            "createdAt": datetime.now().isoformat(timespec="seconds"),
        })
        _write(data)
        return data

    def update_project(self, project_id: str, patch: dict) -> dict:
        data = _read()
        for p in data["projects"]:
            if p["id"] == project_id:
                p.update(patch)
                break
        _write(data)
        return data

    def delete_project(self, project_id: str) -> dict:
        data = _read()
        data["projects"] = [p for p in data["projects"] if p["id"] != project_id]
        _write(data)
        return data

    # ---------- Événements (Calendrier) ----------
    def add_event(self, event: dict) -> dict:
        data = _read()
        data["events"].append({
            "id": str(uuid.uuid4()),
            "calendar": event.get("calendar", "Travail"),
            "title": (event.get("title") or "").strip(),
            "start": event.get("start", ""),
            "end": event.get("end", ""),
            "location": event.get("location", ""),
            "reminder": event.get("reminder", ""),
            "recurrence": event.get("recurrence", "Jamais"),
            "description": event.get("description", ""),
            "createdAt": datetime.now().isoformat(timespec="seconds"),
            "source": event.get("source", "manual"),
        })
        _write(data)
        return data

    def update_event(self, event_id: str, patch: dict) -> dict:
        data = _read()
        for ev in data["events"]:
            if ev["id"] == event_id:
                ev.update(patch)
                break
        _write(data)
        return data

    def delete_event(self, event_id: str) -> dict:
        data = _read()
        # Mémorise les UIDs à supprimer côté iCloud lors de la prochaine sync.
        pending = set(data["icloud"].get("pending_deletes", []))
        for ev in data["events"]:
            if ev["id"] != event_id:
                continue
            if ev.get("source") == "icloud" and ev.get("uid"):
                pending.add(ev["uid"])
            else:
                # Event local potentiellement déjà poussé → UID dérivé de l'id.
                pending.add(f"{ev['id']}@nathancodelist")
            break
        data["icloud"]["pending_deletes"] = sorted(pending)
        data["events"] = [ev for ev in data["events"] if ev["id"] != event_id]
        _write(data)
        return data

    def clear_events(self, source: str = "") -> dict:
        """Supprime tous les events, ou ceux d'une source donnée (ex. 'excel').
        Mémorise les UIDs concernés pour suppression iCloud à la prochaine sync."""
        data = _read()
        pending = set(data["icloud"].get("pending_deletes", []))
        if source:
            removed = [ev for ev in data["events"]
                       if ev.get("source") == source]
            data["events"] = [ev for ev in data["events"]
                              if ev.get("source") != source]
        else:
            removed = list(data["events"])
            data["events"] = []
        for ev in removed:
            if ev.get("source") == "icloud" and ev.get("uid"):
                pending.add(ev["uid"])
            else:
                pending.add(f"{ev['id']}@nathancodelist")
        data["icloud"]["pending_deletes"] = sorted(pending)
        _write(data)
        return data

    # ---------- Sync iCloud (CalDAV, push one-way) ----------
    def save_icloud_credentials(self, apple_id: str, app_password: str,
                                calendar_name: str = "Nathan Code List") -> dict:
        data = _read()
        data["icloud"]["apple_id"] = (apple_id or "").strip()
        data["icloud"]["app_password"] = (app_password or "").strip()
        data["icloud"]["calendar_name"] = (
            calendar_name or "Nathan Code List").strip()
        _write(data)
        return {"ok": True}

    def get_icloud_status(self) -> dict:
        data = _read()
        ic = data["icloud"]
        return {
            "configured": bool(ic.get("apple_id") and ic.get("app_password")),
            "apple_id": ic.get("apple_id", ""),
            "calendar_name": ic.get("calendar_name", "Nathan Code List"),
            "last_sync": ic.get("last_sync", ""),
        }

    def test_icloud_connection(self) -> dict:
        if caldav is None:
            return {"ok": False, "error": "Module caldav non installé."}
        data = _read()
        ic = data["icloud"]
        if not ic.get("apple_id") or not ic.get("app_password"):
            return {"ok": False, "error": "Identifiants iCloud manquants."}
        try:
            client = caldav.DAVClient(
                url=ICLOUD_CALDAV_URL,
                username=ic["apple_id"],
                password=ic["app_password"],
            )
            principal = client.principal()
            calendars = [c.name or "(sans nom)" for c in principal.calendars()]
            return {"ok": True, "calendars": calendars}
        except AuthorizationError:
            return {"ok": False, "error":
                    "Authentification refusée. Vérifie ton Apple ID et "
                    "utilise un mot de passe d'app (appleid.apple.com)."}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def sync_to_icloud(self) -> dict:
        """Sync bidirectionnelle avec le calendrier iCloud dédié :
         - Pull : tous les events distants → locaux (source 'icloud')
         - Push : tous les events locaux non-icloud → iCloud (create-or-update
           via UID stable). Les events ayant disparu d'iCloud sont supprimés
           localement (côté source='icloud' uniquement).
        """
        if caldav is None:
            return {"ok": False, "error": "Module caldav non installé."}
        data = _read()
        ic = data["icloud"]
        if not ic.get("apple_id") or not ic.get("app_password"):
            return {"ok": False, "error":
                    "Configure d'abord tes identifiants iCloud."}

        try:
            client = caldav.DAVClient(
                url=ICLOUD_CALDAV_URL,
                username=ic["apple_id"],
                password=ic["app_password"],
            )
            principal = client.principal()
            target_name = ic.get("calendar_name") or "Nathan Code List"

            target = None
            for c in principal.calendars():
                if (c.name or "") == target_name:
                    target = c
                    break
            if target is None:
                try:
                    target = principal.make_calendar(name=target_name)
                except Exception as exc:
                    return {"ok": False, "error":
                            f"Impossible de créer le calendrier iCloud : {exc}"}

            # --- DELETE : appliquer les suppressions locales sur iCloud ---
            # Stratégie : on itère TOUS les events de TOUS les calendriers
            # et on supprime ceux dont l'UID est dans la liste tombstone.
            # Plus robuste que event_by_uid qui échoue parfois silencieusement
            # avec iCloud.
            pending = set(data["icloud"].get("pending_deletes", []))
            deleted_remote = 0
            deleted_log = []
            try:
                all_cals_for_delete = list(principal.calendars())
            except Exception:
                all_cals_for_delete = [target]
            if pending:
                for cal in all_cals_for_delete:
                    cal_name = cal.name or "iCloud"
                    try:
                        cal_events = list(cal.events())
                    except Exception:
                        continue
                    for cal_event in cal_events:
                        try:
                            ics = cal_event.data
                        except Exception:
                            continue
                        parsed = _ics_to_event(ics, fallback_calendar=cal_name)
                        if not parsed:
                            continue
                        if parsed["uid"] not in pending:
                            continue
                        try:
                            cal_event.delete()
                            deleted_remote += 1
                            deleted_log.append(
                                f"{parsed['title']} ({cal_name})")
                        except Exception as exc:
                            deleted_log.append(
                                f"ERREUR {parsed['title']}: {exc}")
            # Tombstones purgées dans tous les cas : si un UID n'existe pas
            # côté iCloud, c'est qu'il a déjà été supprimé.
            data["icloud"]["pending_deletes"] = []
            deleted_uids = set(pending)

            # --- PUSH : envoyer les events locaux non-icloud ---
            pushed, push_errors = 0, []
            for ev in data["events"]:
                if ev.get("source") == "icloud":
                    continue
                ics = _event_to_ics(ev)
                if not ics:
                    continue
                uid = f"{ev.get('id')}@nathancodelist"
                try:
                    existing = None
                    try:
                        existing = target.event_by_uid(uid)
                    except NotFoundError:
                        existing = None
                    except Exception:
                        existing = None
                    if existing is not None:
                        existing.data = ics
                        existing.save()
                    else:
                        target.save_event(ics)
                    pushed += 1
                except Exception as exc:
                    push_errors.append(f"{ev.get('title','?')}: {exc}")

            # --- PULL : récupérer les events de TOUS les calendriers iCloud
            # (pas seulement le calendrier "target" — sinon les events ajoutés
            # sur iPhone dans Perso/Maison/etc. ne reviennent pas).
            remote_events = []
            seen_uids = set()
            pulled_from = []
            try:
                all_calendars = list(principal.calendars())
            except Exception as exc:
                return {"ok": False,
                        "error": f"Lecture des calendriers iCloud échouée : {exc}"}

            for cal in all_calendars:
                cal_name = cal.name or "iCloud"
                try:
                    cal_events = list(cal.events())
                except Exception:
                    continue
                pulled_from.append(f"{cal_name} ({len(cal_events)})")
                for cal_event in cal_events:
                    try:
                        ics = cal_event.data
                    except Exception:
                        continue
                    parsed = _ics_to_event(ics, fallback_calendar=cal_name)
                    if not parsed:
                        continue
                    if parsed["uid"] in seen_uids:
                        continue
                    if parsed["uid"] in deleted_uids:
                        # On vient de supprimer cet UID dans la même sync ;
                        # iCloud peut encore le retourner en cache → on l'ignore.
                        continue
                    seen_uids.add(parsed["uid"])
                    remote_events.append(parsed)

            # On garde les events locaux non-icloud tels quels, et on remplace
            # le sous-ensemble source='icloud' par le miroir d'iCloud.
            non_icloud = [ev for ev in data["events"]
                          if ev.get("source") != "icloud"]
            # Pour éviter d'afficher 2x un event poussé puis re-pull :
            # on ignore les events distants dont l'UID correspond à un event
            # local non-icloud (`<id>@nathancodelist`).
            local_pushed_uids = {f"{ev.get('id')}@nathancodelist"
                                 for ev in non_icloud}
            new_icloud = []
            for r in remote_events:
                if r["uid"] in local_pushed_uids:
                    continue
                new_icloud.append({
                    "id": str(uuid.uuid4()),
                    "uid": r["uid"],
                    "calendar": r["calendar"],
                    "title": r["title"],
                    "start": r["start"],
                    "end": r["end"],
                    "location": r["location"],
                    "reminder": r["reminder"],
                    "recurrence": r.get("recurrence", "Jamais"),
                    "description": r["description"],
                    "createdAt": datetime.now().isoformat(timespec="seconds"),
                    "source": "icloud",
                })

            data["events"] = non_icloud + new_icloud
            data["icloud"]["last_sync"] = datetime.now().isoformat(
                timespec="seconds")
            _write(data)
            return {
                "ok": True,
                "pushed": pushed,
                "pulled": len(new_icloud),
                "deleted": deleted_remote,
                "deleted_log": deleted_log,
                "tombstones_processed": len(deleted_uids),
                "errors": push_errors,
                "last_sync": data["icloud"]["last_sync"],
                "calendars_scanned": pulled_from,
                "data": data,
            }

        except AuthorizationError:
            return {"ok": False, "error":
                    "Authentification refusée. Régénère un mot de passe d'app."}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def import_events_from_excel(self, replace_excel: bool = False) -> dict:
        """Ouvre un dialogue natif, parse le .xlsx, ajoute les events.
        Si `replace_excel=True`, remplace d'abord tous les events 'excel'."""
        win = webview.windows[0] if webview.windows else None
        if win is None:
            return {"error": "Fenêtre indisponible", "imported": 0}
        file_types = ("Fichiers Excel (*.xlsx;*.xlsm)", "Tous les fichiers (*.*)")
        paths = win.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types
        )
        if not paths:
            return {"cancelled": True, "imported": 0}
        path = paths[0]
        try:
            new_events = _parse_events_from_xlsx(path)
        except Exception as exc:
            return {"error": str(exc), "imported": 0}

        data = _read()
        if replace_excel:
            data["events"] = [ev for ev in data["events"]
                              if ev.get("source") != "excel"]
        data["events"].extend(new_events)
        _write(data)
        return {"imported": len(new_events), "data": data, "path": path}


# ----------------------------------------------------------------------
# Bootstrap : crée une vraie fenêtre native (WebView2 sur Windows 11)
# ----------------------------------------------------------------------
def main() -> None:
    _ensure_data_file()

    webview.create_window(
        title="Nathan Code List",
        url=INDEX_FILE,
        js_api=Api(),
        width=1400,
        height=900,
        x=80,
        y=40,
        min_size=(1000, 700),
        background_color="#E0F3E1",  # cohérent avec le light mode au boot
    )
    # gui="edgechromium" → WebView2 (livré avec Windows 10/11, aucune install)
    # debug=True → DevTools accessibles via F12 / clic droit → Inspecter
    webview.start(gui="edgechromium", debug=True)


if __name__ == "__main__":
    main()
